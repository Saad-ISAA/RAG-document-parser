# parsers/spreadsheet_parser.py
"""
Spreadsheet parser for Excel (XLSX, XLS), OpenDocument (ODS), and CSV files.
"""

import logging
from pathlib import Path
from typing import List, Optional, Dict, Any, Union
import csv
from datetime import datetime

# Import spreadsheet libraries with fallbacks
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
    pd = None

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    openpyxl = None

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False
    xlrd = None

try:
    from odf import table, text
    from odf.opendocument import load as odf_load
    HAS_ODFPY = True
except ImportError:
    HAS_ODFPY = False

from models.parse_result import ParseResult, TableData, DocumentMetadata, FileInfo
from utils.config import ParserConfig

logger = logging.getLogger(__name__)


class SpreadsheetParser:
    """
    Parser for spreadsheet formats including XLSX, XLS, ODS, and CSV.
    """
    
    def __init__(self, config: ParserConfig):
        """Initialize spreadsheet parser with configuration."""
        self.config = config
        self.sheet_config = config.spreadsheet
        
        # Check available libraries
        self.available_libs = {}
        if HAS_PANDAS:
            self.available_libs['pandas'] = True
        if HAS_OPENPYXL:
            self.available_libs['openpyxl'] = True
        if HAS_XLRD:
            self.available_libs['xlrd'] = True
        if HAS_ODFPY:
            self.available_libs['odfpy'] = True
        
        logger.info(f"Spreadsheet parser initialized with libraries: {', '.join(self.available_libs.keys())}")
    
    def parse(self, file_path: Path, file_info: FileInfo) -> ParseResult:
        """
        Parse a spreadsheet file.
        
        Args:
            file_path: Path to the spreadsheet file
            file_info: File information from detector
            
        Returns:
            ParseResult with extracted content and tables
        """
        result = ParseResult(
            file_path=str(file_path),
            file_type=file_info.mime_type
        )
        
        extension = file_path.suffix.lower()
        
        try:
            if extension == '.csv':
                return self._parse_csv(file_path, result)
            elif extension == '.xlsx':
                return self._parse_xlsx(file_path, result)
            elif extension == '.xls':
                return self._parse_xls(file_path, result)
            elif extension == '.ods':
                return self._parse_ods(file_path, result)
            else:
                result.error = f"Unsupported spreadsheet format: {extension}"
                return result
                
        except Exception as e:
            logger.error(f"Error parsing spreadsheet {file_path}: {str(e)}")
            result.error = str(e)
            return result
    
    def _parse_csv(self, file_path: Path, result: ParseResult) -> ParseResult:
        """Parse CSV file."""
        try:
            # Detect CSV dialect
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                sample = f.read(1024)
                sniffer = csv.Sniffer()
                dialect = sniffer.sniff(sample)
            
            # Read CSV data
            rows = []
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f, dialect=dialect)
                for row in reader:
                    if len(rows) >= self.sheet_config.max_rows:
                        break
                    
                    # Limit columns
                    if len(row) > self.sheet_config.max_columns:
                        row = row[:self.sheet_config.max_columns]
                    
                    # Skip empty rows if configured
                    if self.sheet_config.skip_empty_rows and not any(cell.strip() for cell in row):
                        continue
                    
                    rows.append(row)
            
            if not rows:
                result.error = "No data found in CSV file"
                return result
            
            # Create table data
            headers = rows[0] if rows else []
            data_rows = rows[1:] if len(rows) > 1 else []
            
            # Create table
            table = TableData(
                headers=headers,
                rows=data_rows,
                table_index=0
            )
            result.tables = [table]
            
            # Create text content
            result.content = self._create_text_from_table(table)
            
            # Create metadata
            result.metadata = self._create_csv_metadata(rows, file_path)
            
            result.success = True
            result.parser_used = 'csv'
            
        except Exception as e:
            result.error = f"CSV parsing failed: {str(e)}"
        
        return result
    
    def _parse_xlsx(self, file_path: Path, result: ParseResult) -> ParseResult:
        """Parse XLSX file using openpyxl or pandas."""
        if HAS_PANDAS:
            return self._parse_xlsx_pandas(file_path, result)
        elif HAS_OPENPYXL:
            return self._parse_xlsx_openpyxl(file_path, result)
        else:
            result.error = "No XLSX parsing library available (install pandas or openpyxl)"
            return result
    
    def _parse_xlsx_pandas(self, file_path: Path, result: ParseResult) -> ParseResult:
        """Parse XLSX using pandas."""
        try:
            # Read Excel file
            if self.sheet_config.read_all_sheets:
                excel_data = pd.read_excel(file_path, sheet_name=None, nrows=self.sheet_config.max_rows)
                sheets_data = excel_data
            else:
                # Read only first sheet
                df = pd.read_excel(file_path, nrows=self.sheet_config.max_rows)
                sheets_data = {'Sheet1': df}
            
            # Process each sheet
            all_content = []
            for sheet_name, df in sheets_data.items():
                # Limit columns
                if len(df.columns) > self.sheet_config.max_columns:
                    df = df.iloc[:, :self.sheet_config.max_columns]
                
                # Skip empty rows/columns if configured
                if self.sheet_config.skip_empty_rows:
                    df = df.dropna(how='all')
                if self.sheet_config.skip_empty_columns:
                    df = df.dropna(how='all', axis=1)
                
                # Convert to table data
                headers = [str(col) for col in df.columns]
                rows = []
                for _, row in df.iterrows():
                    row_data = [str(val) if pd.notna(val) else "" for val in row]
                    rows.append(row_data)
                
                if headers or rows:
                    table = TableData(
                        headers=headers,
                        rows=rows,
                        table_index=len(result.tables)
                    )
                    result.tables.append(table)
                    
                    # Add to content
                    sheet_content = f"=== Sheet: {sheet_name} ===\n"
                    sheet_content += self._create_text_from_table(table)
                    all_content.append(sheet_content)
            
            result.content = '\n\n'.join(all_content)
            result.metadata = self._create_excel_metadata(file_path)
            result.success = True
            result.parser_used = 'pandas'
            
        except Exception as e:
            result.error = f"XLSX parsing with pandas failed: {str(e)}"
        
        return result
    
    def _parse_xlsx_openpyxl(self, file_path: Path, result: ParseResult) -> ParseResult:
        """Parse XLSX using openpyxl."""
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            # Get sheets to process
            if self.sheet_config.read_all_sheets:
                sheet_names = workbook.sheetnames
            else:
                sheet_names = [workbook.sheetnames[0]] if workbook.sheetnames else []
            
            all_content = []
            for sheet_name in sheet_names:
                sheet = workbook[sheet_name]
                
                # Extract data
                rows_data = []
                for row_num, row in enumerate(sheet.iter_rows(max_row=self.sheet_config.max_rows), 1):
                    if len(row) > self.sheet_config.max_columns:
                        row = row[:self.sheet_config.max_columns]
                    
                    row_data = []
                    for cell in row:
                        value = cell.value
                        if value is None:
                            row_data.append("")
                        else:
                            row_data.append(str(value))
                    
                    # Skip empty rows if configured
                    if self.sheet_config.skip_empty_rows and not any(cell.strip() for cell in row_data):
                        continue
                    
                    rows_data.append(row_data)
                
                if rows_data:
                    # First row as headers
                    headers = rows_data[0] if rows_data else []
                    data_rows = rows_data[1:] if len(rows_data) > 1 else []
                    
                    table = TableData(
                        headers=headers,
                        rows=data_rows,
                        table_index=len(result.tables)
                    )
                    result.tables.append(table)
                    
                    # Add to content
                    sheet_content = f"=== Sheet: {sheet_name} ===\n"
                    sheet_content += self._create_text_from_table(table)
                    all_content.append(sheet_content)
            
            result.content = '\n\n'.join(all_content)
            result.metadata = self._create_excel_metadata(file_path)
            result.success = True
            result.parser_used = 'openpyxl'
            
        except Exception as e:
            result.error = f"XLSX parsing with openpyxl failed: {str(e)}"
        
        return result
    
    def _parse_xls(self, file_path: Path, result: ParseResult) -> ParseResult:
        """Parse XLS file using xlrd or pandas."""
        if HAS_PANDAS:
            try:
                # Try pandas first
                if self.sheet_config.read_all_sheets:
                    excel_data = pd.read_excel(file_path, sheet_name=None, nrows=self.sheet_config.max_rows)
                    sheets_data = excel_data
                else:
                    df = pd.read_excel(file_path, nrows=self.sheet_config.max_rows)
                    sheets_data = {'Sheet1': df}
                
                # Process similar to XLSX
                all_content = []
                for sheet_name, df in sheets_data.items():
                    if len(df.columns) > self.sheet_config.max_columns:
                        df = df.iloc[:, :self.sheet_config.max_columns]
                    
                    if self.sheet_config.skip_empty_rows:
                        df = df.dropna(how='all')
                    if self.sheet_config.skip_empty_columns:
                        df = df.dropna(how='all', axis=1)
                    
                    headers = [str(col) for col in df.columns]
                    rows = []
                    for _, row in df.iterrows():
                        row_data = [str(val) if pd.notna(val) else "" for val in row]
                        rows.append(row_data)
                    
                    if headers or rows:
                        table = TableData(
                            headers=headers,
                            rows=rows,
                            table_index=len(result.tables)
                        )
                        result.tables.append(table)
                        
                        sheet_content = f"=== Sheet: {sheet_name} ===\n"
                        sheet_content += self._create_text_from_table(table)
                        all_content.append(sheet_content)
                
                result.content = '\n\n'.join(all_content)
                result.metadata = self._create_excel_metadata(file_path)
                result.success = True
                result.parser_used = 'pandas'
                
            except Exception as e:
                result.error = f"XLS parsing with pandas failed: {str(e)}"
        
        elif HAS_XLRD:
            result.error = "XLS parsing with xlrd not fully implemented (use pandas instead)"
        else:
            result.error = "No XLS parsing library available (install pandas)"
        
        return result
    
    def _parse_ods(self, file_path: Path, result: ParseResult) -> ParseResult:
        """Parse ODS file using odfpy."""
        if not HAS_ODFPY:
            result.error = "odfpy library not available for ODS parsing"
            return result
        
        try:
            doc = odf_load(file_path)
            
            # Get all tables (sheets)
            tables = doc.getElementsByType(table.Table)
            
            all_content = []
            for table_index, ods_table in enumerate(tables):
                table_name = ods_table.getAttribute('name') or f"Sheet{table_index + 1}"
                
                # Extract rows
                rows_data = []
                table_rows = ods_table.getElementsByType(table.TableRow)
                
                for row_elem in table_rows[:self.sheet_config.max_rows]:
                    cells = row_elem.getElementsByType(table.TableCell)
                    row_data = []
                    
                    for cell in cells[:self.sheet_config.max_columns]:
                        # Get cell text content
                        cell_text = ""
                        for p in cell.getElementsByType(text.P):
                            if p.firstChild:
                                cell_text += str(p.firstChild)
                        row_data.append(cell_text)
                    
                    # Skip empty rows if configured
                    if self.sheet_config.skip_empty_rows and not any(cell.strip() for cell in row_data):
                        continue
                    
                    rows_data.append(row_data)
                
                if rows_data:
                    # First row as headers
                    headers = rows_data[0] if rows_data else []
                    data_rows = rows_data[1:] if len(rows_data) > 1 else []
                    
                    table_data = TableData(
                        headers=headers,
                        rows=data_rows,
                        table_index=table_index
                    )
                    result.tables.append(table_data)
                    
                    # Add to content
                    sheet_content = f"=== Sheet: {table_name} ===\n"
                    sheet_content += self._create_text_from_table(table_data)
                    all_content.append(sheet_content)
                
                if not self.sheet_config.read_all_sheets:
                    break  # Only process first sheet
            
            result.content = '\n\n'.join(all_content)
            result.metadata = self._create_ods_metadata(file_path)
            result.success = True
            result.parser_used = 'odfpy'
            
        except Exception as e:
            result.error = f"ODS parsing failed: {str(e)}"
        
        return result
    
    def _create_text_from_table(self, table: TableData) -> str:
        """Convert table data to readable text format."""
        if not table.headers and not table.rows:
            return ""
        
        lines = []
        
        # Add headers
        if table.headers:
            lines.append(" | ".join(table.headers))
            lines.append("-" * len(lines[0]))
        
        # Add data rows
        for row in table.rows:
            lines.append(" | ".join(row))
        
        return "\n".join(lines)
    
    def _create_csv_metadata(self, rows: List[List[str]], file_path: Path) -> DocumentMetadata:
        """Create metadata for CSV file."""
        metadata = DocumentMetadata()
        
        if rows:
            metadata.word_count = sum(len(cell.split()) for row in rows for cell in row)
            metadata.page_count = 1
        
        # File stats
        try:
            stat = file_path.stat()
            metadata.modification_date = datetime.fromtimestamp(stat.st_mtime)
        except Exception:
            pass
        
        # Custom properties
        metadata.custom_properties = {
            'row_count': len(rows),
            'column_count': len(rows[0]) if rows else 0,
            'file_format': 'CSV'
        }
        
        return metadata
    
    def _create_excel_metadata(self, file_path: Path) -> DocumentMetadata:
        """Create metadata for Excel file."""
        metadata = DocumentMetadata()
        
        # File stats
        try:
            stat = file_path.stat()
            metadata.modification_date = datetime.fromtimestamp(stat.st_mtime)
        except Exception:
            pass
        
        # Try to extract built-in properties if using openpyxl
        if HAS_OPENPYXL:
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                props = wb.properties
                
                metadata.title = props.title
                metadata.author = props.creator
                metadata.subject = props.subject
                metadata.creation_date = props.created
                metadata.modification_date = props.modified
                
            except Exception as e:
                logger.debug(f"Failed to extract Excel metadata: {str(e)}")
        
        metadata.custom_properties = {
            'file_format': 'Excel',
            'extension': file_path.suffix.upper()
        }
        
        return metadata
    
    def _create_ods_metadata(self, file_path: Path) -> DocumentMetadata:
        """Create metadata for ODS file."""
        metadata = DocumentMetadata()
        
        # File stats
        try:
            stat = file_path.stat()
            metadata.modification_date = datetime.fromtimestamp(stat.st_mtime)
        except Exception:
            pass
        
        metadata.custom_properties = {
            'file_format': 'OpenDocument Spreadsheet'
        }
        
        return metadata
    
    def get_supported_formats(self) -> List[str]:
        """Get list of supported spreadsheet formats."""
        formats = ['csv']
        
        if HAS_PANDAS or HAS_OPENPYXL:
            formats.append('xlsx')
        if HAS_PANDAS or HAS_XLRD:
            formats.append('xls')
        if HAS_ODFPY:
            formats.append('ods')
        
        return formats
    
    def get_library_info(self) -> Dict[str, Any]:
        """Get information about available spreadsheet libraries."""
        return {
            'available_libraries': self.available_libs,
            'capabilities': {
                'csv': {
                    'reading': True,
                    'dialect_detection': True,
                    'encoding_handling': True
                },
                'xlsx': {
                    'reading': HAS_PANDAS or HAS_OPENPYXL,
                    'multiple_sheets': HAS_PANDAS or HAS_OPENPYXL,
                    'metadata_extraction': HAS_OPENPYXL
                },
                'xls': {
                    'reading': HAS_PANDAS,
                    'multiple_sheets': HAS_PANDAS,
                    'metadata_extraction': False
                },
                'ods': {
                    'reading': HAS_ODFPY,
                    'multiple_sheets': HAS_ODFPY,
                    'metadata_extraction': False
                }
            }
        }
    
    def analyze_spreadsheet_structure(self, file_path: Path) -> Dict[str, Any]:
        """Analyze spreadsheet structure without full parsing."""
        try:
            extension = file_path.suffix.lower()
            
            if extension == '.csv':
                return self._analyze_csv_structure(file_path)
            elif extension == '.xlsx' and HAS_OPENPYXL:
                return self._analyze_xlsx_structure(file_path)
            else:
                return {'error': 'Structure analysis not supported for this format'}
                
        except Exception as e:
            return {'error': str(e)}
    
    def _analyze_csv_structure(self, file_path: Path) -> Dict[str, Any]:
        """Analyze CSV file structure."""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                # Read first few lines to analyze structure
                lines = []
                for i, line in enumerate(f):
                    lines.append(line.strip())
                    if i >= 10:  # First 10 lines should be enough
                        break
            
            if not lines:
                return {'rows': 0, 'columns': 0}
            
            # Detect dialect
            sniffer = csv.Sniffer()
            dialect = sniffer.sniff('\n'.join(lines))
            
            # Count columns in first row
            reader = csv.reader([lines[0]], dialect=dialect)
            first_row = next(reader, [])
            
            return {
                'estimated_rows': len(lines),
                'columns': len(first_row),
                'delimiter': dialect.delimiter,
                'has_header': sniffer.has_header('\n'.join(lines)),
                'sample_data': lines[:3]
            }
            
        except Exception as e:
            return {'error': str(e)}
    
    def _analyze_xlsx_structure(self, file_path: Path) -> Dict[str, Any]:
        """Analyze XLSX file structure."""
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            
            sheets_info = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Get sheet dimensions
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                sheets_info.append({
                    'name': sheet_name,
                    'rows': max_row,
                    'columns': max_col,
                    'cells_with_data': sum(1 for row in sheet.iter_rows() for cell in row if cell.value)
                })
            
            return {
                'sheet_count': len(wb.sheetnames),
                'sheets': sheets_info
            }
            
        except Exception as e:
            return {'error': str(e)}